import shutil
import os
from datetime import datetime
import sqlalchemy as sa
import sqlalchemy_access as sa_a
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, colors
from dotenv import load_dotenv


def backup_database(db_path: str, backup_dir: str = None) -> str:
    """
    Creates a backup copy of the Access database file before any modifications are made.

    Args:
        db_path (str): The full path to the Access database file (.mdb or .accdb).
        backup_dir (str, optional): The directory where the backup will be stored. 
                                    If None, the backup will be stored in the same directory as the original file.
    
    Returns:
        str: The full path to the backup file.
    
    Raises:
        FileNotFoundError: If the original database file does not exist.
    
    Example:
        backup_file = backup_database('C:/path/to/database.mdb')
    """
    
    # Check if the database file exists
    if not os.path.exists(db_path):
        raise FileNotFoundError(f"The database file {db_path} does not exist.")
    
    # Get the current timestamp to append to the backup file name
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Get the directory and file name of the database
    db_dir, db_name = os.path.split(db_path)
    db_base, db_ext = os.path.splitext(db_name)
    
    # If no backup directory is provided, use the same directory as the database
    if backup_dir is None:
        backup_dir = db_dir
    
    # Create the backup file name with the timestamp
    backup_filename = f"{db_base}_backup_{timestamp}{db_ext}"
    backup_path = os.path.join(backup_dir, backup_filename)
    
    # Perform the file copy to create the backup
    shutil.copy2(db_path, backup_path)
    
    return backup_path


def get_dbaccess_connection(db_path: str):
    """
    Establishes a connection to a Microsoft Access database using SQLAlchemy 
    and returns a connection engine.

    Args:
        db_path (str): The full path to the Access database file (.mdb or .accdb).

    Returns:
        sqlalchemy.engine.Engine: An SQLAlchemy engine object representing the connection 
        to the Access database.

    Details:
        - Extracts the database name from the provided `db_path`.
        - Uses the ODBC driver "{Microsoft Access Driver (*.mdb, *.accdb)}" to establish the connection.
        - Configures the connection with the "ExtendedAnsiSQL=1" parameter for better ANSI compatibility.
        - Leverages `sqlalchemy_access` and `pyodbc` for SQLAlchemy integration.

    Example:
        engine = get_dbaccess_connection('C:/path/to/your_database.mdb')
    
    """
    # Extract the database name from the db_path
    db_name = os.path.splitext(os.path.basename(db_path))[0]

    #
    # Create the SQLAlchemy engine with an ODBC connection string
    # Ordinary unprotected Access database
    #  
    driver = "{Microsoft Access Driver (*.mdb, *.accdb)}"
    connection_string = (
        f"DRIVER={driver};"
        f"DBQ={db_path};"
        f"ExtendedAnsiSQL=1;"
        )
    connection_url = sa.engine.URL.create(
        "access+pyodbc",
        query={"odbc_connect": connection_string}
        )
    engine = sa.create_engine(connection_url)

    return engine


def get_database_name_from_path(db_path: str) -> str:
    """
    Extracts the database name from a full file path.
    
    Args:
        db_path (str): The full path to the database file.
    
    Returns:
        str: The name of the database without the file extension.
    """
    # Extract the file name from the path
    db_file = os.path.basename(db_path)
    
    # Remove the file extension and return the name
    db_name = os.path.splitext(db_file)[0]
    
    return db_name


def has_trailing_letter(code) -> bool:
    """
    Checks if a given string has a letter (A-Z or a-z) at the end.

    Args:
        code (str): The string to be checked.

    Returns:
        bool: True if the string ends with a letter, False otherwise.

    Example:
        has_trailing_letter("12345A")  # Returns True
        has_trailing_letter("123456")  # Returns False
    """
    return bool(re.match(r'.*[A-Za-z]$', code))


def generate_unique_code(old_code, existing_codes) -> str:
    """
    Generates a new unique code by removing the last letter from the old code 
    and prefixing it with a number (starting with '9'). If the generated code 
    already exists, it tries successive prefixes ('8', '7', etc.) until a unique 
    code is found.

    Args:
        old_code (str): The original code containing a trailing letter.
        existing_codes (set): A set of codes that are already in use.

    Returns:
        str: A new unique code that is not present in `existing_codes`.

    Raises:
        ValueError: If a unique code could not be generated after trying all prefixes.

    Example:
        generate_unique_code("12345A", {"912345", "812345"})  # Returns a unique code like "712345"
    """
    base_code = old_code[:-1]  # Remove the trailing letter
    for prefix in ['9', '8', '7', '6', '5', '4', '3', '2', '1', '0']:
        new_code = prefix + base_code
        if new_code not in existing_codes:
            return new_code
    raise ValueError("Could not generate a unique code")


def format_header_cell(cell, font_size=11):
    """
    Formats a header cell with the default styling: white bold text and green background.

    Parameters:
    -----------
    cell : openpyxl.cell.cell.Cell
        The cell to format.
    
    font_size : int, optional
        The font size to be applied to the header cell. Default is 11.

    Returns:
    --------
    None
    """
    cell.font = Font(color="FFFFFF", bold=True, size=font_size + 1)
    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")


def adjust_column_widths(sheet,  max_width=80):
    """
    Adjusts the width of each column in the Excel sheet based on the maximum width of the data and header values.

    Parameters:
    -----------
    sheet : openpyxl.worksheet.worksheet.Worksheet
        The worksheet where column widths need to be adjusted.

    max_width : int, optional (default=80)
        The maximum allowed width for any column. If the calculated width exceeds this value,
        the column width will be set to this maximum value.
    
    Returns:
    --------
    None
    """
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name

        # Calculate the width required by the header (considering formatting)
        header_length = len(str(col[0].value))
        adjusted_header_length = header_length * 1.5  # Factor to account for bold and larger font size

        # Compare the header length with the lengths of the data values
        for cell in col:
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        
        # Use the greater of the header length or data length for column width
        max_length = max(max_length, adjusted_header_length)

        # Adjust the column width and apply the max_width limit
        adjusted_width = min(max_length + 3, max_width)
        sheet.column_dimensions[column].width = adjusted_width


def save_and_format_dataframe_to_excel(dfs_dict: dict, excel_path: str):
    """
    Saves a dictionary of DataFrames to an Excel file, with each DataFrame saved on a separate sheet.
    Formats the headers, adjusts column widths, and applies filters to all columns for each sheet.

    Args:
        dfs_dict (dict): A dictionary where keys are sheet names and values are DataFrames to be saved.
        excel_path (str): The file path where the Excel file will be saved.
    """
    # Create a new Excel writer object
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        for sheet_title, df in dfs_dict.items():
            # Write each DataFrame to a separate sheet
            df.to_excel(writer, index=False, sheet_name=sheet_title)
    
    # Load the workbook to apply formatting
    wb = load_workbook(excel_path)
    
    # Loop through each sheet in the workbook
    for sheet_title in dfs_dict.keys():
        sheet = wb[sheet_title]  # Access the sheet by name
        
        # Format header cells for all columns
        for col_num in range(1, dfs_dict[sheet_title].shape[1] + 1):  # Loop over the number of columns
            format_header_cell(sheet.cell(row=1, column=col_num))  # Apply header formatting

        # Call the adjust_column_widths function to format the sheet
        adjust_column_widths(sheet)

        # Apply a filter to all columns
        sheet.auto_filter.ref = sheet.dimensions

    # Save the formatted Excel file
    wb.save(excel_path)


def get_db_keys_from_env(database_name: str):
    """
    Extracts table key fields from environment variables that follow the format:
    <database>__<table>=<list of key fields separated by commas>.
    
    Args:
        database_name (str): The name of the database to filter the keys.
    
    Returns:
        dict: A dictionary where the keys are table names and the values are lists of key fields.
    """
    # Load environment variables from .env file
    load_dotenv()

    # Initialize an empty dictionary to store table keys
    db_keys = {}

    # Iterate over environment variables
    for key, value in os.environ.items():
        # Check if the key starts with the database name followed by '__'
        if key.startswith(f"{database_name}__"):
            # Extract the table name by removing the database name and '__'
            table_name = key.split('__')[1]
            
            # Split the value into a list of key fields, assuming they are separated by commas
            key_fields = [field.strip() for field in value.split(',')]
            
            # Add the table name and key fields to the dictionary
            db_keys[table_name] = key_fields

    return db_keys


def find_code_matches_in_db(df_mappings: pd.DataFrame, db_path: str, db_keys: dict) -> dict:
    """
    Search all alphanumeric fields in all non-empty tables in the database for old client codes 
    that need to be replaced. The fields must have a length between 6 and 20 characters. Matches 
    are returned for each table as a DataFrame containing the key columns and the columns where 
    the match was found. Also adds columns for FOUND_FIELD and NEW_VALUE to prepare for updates.

    Args:
        df_mappings (pd.DataFrame): A DataFrame containing the old and new codes.
        db_path (str): Path to the Access database.
        db_keys (dict): Dictionary containing key fields for each table.
    
    Returns:
        dict: A dictionary where keys are table names, and values are DataFrames with the matches.
              Each DataFrame contains the key fields (if available), the column where the match 
              was found, the matched value, and the corresponding new value.
    """
    # Connect to the database
    engine = get_dbaccess_connection(db_path)

    # Initialize an empty dictionary to store matches for each table
    matches_dict = {}
    
    # Get the list of all tables in the database
    inspector = sa.inspect(engine)
    tables = inspector.get_table_names()

    # Loop through each table in the database
    for table_name in tables:
        # Check if the table has any records by querying the count of rows
        query_count = f"SELECT COUNT(*) AS countregs FROM {table_name}"
        try:
            result = pd.read_sql(query_count, engine)
            if result['countregs'][0] == 0:
                # If the table is empty, skip to the next table
                continue
        except Exception as e:
            print(f"Error counting rows in {table_name}: {e}")
            continue
        
        # Get the columns of the current table
        columns = inspector.get_columns(table_name)
        
        # List to store the columns that match the criteria
        matching_columns = []

        # Loop through each column to find alphanumeric fields with length between 6 and 20
        for column in columns:
            col_name = column['name']
            col_type = str(column['type']).upper()

            # Check if the column is of alphanumeric type and has the correct length
            if ('CHAR' in col_type or 'TEXT' in col_type or 'VARCHAR' in col_type):
                col_length = column['type'].length  # Get the column length

                if col_length and 6 <= col_length <= 20:
                    matching_columns.append(col_name)

        # If no columns match the criteria, skip the table
        if not matching_columns:
            continue

        # Retrieve the key columns for the table from db_keys
        key_columns = db_keys.get(table_name, [])

        # Remove key columns from matching_columns to avoid duplicates
        matching_columns = [col for col in matching_columns if col not in key_columns]

        # Construct the SQL query
        if key_columns:
            key_columns_str = ', '.join(key_columns)
            query = f"SELECT {key_columns_str}, {', '.join(matching_columns)} FROM {table_name}"
        else:
            continue
        
        try:
            df_table = pd.read_sql(query, engine)
        except Exception as e:
            print(f"Error retrieving records from {table_name}: {e}")
            continue

        # Initialize an empty DataFrame to store matches for this table
        df_table_matches = pd.DataFrame()

        # Loop through the old codes and check if they exist in any of the matching columns
        for old_code in df_mappings['OLD_CODE']:
            # Get the corresponding new code for the old code from df_mappings
            new_code = df_mappings.loc[df_mappings['OLD_CODE'] == old_code, 'NEW_CODE'].values[0]

            # Loop through each matching column and check for the old code
            for col_name in matching_columns:
                # Create a copy to avoid SettingWithCopyWarning
                df_code_matches = df_table[df_table[col_name] == old_code].copy()

                # If matches are found, add them to the DataFrame for this table
                if not df_code_matches.empty:
                    df_code_matches.loc[:, 'FOUND_VALUE'] = old_code  # Store the old code found
                    df_code_matches.loc[:, 'FOUND_FIELD'] = col_name  # Store the column where the match was found
                    df_code_matches.loc[:, 'NEW_VALUE'] = new_code    # Store the new code for the update
                    df_table_matches = pd.concat([df_table_matches, df_code_matches], ignore_index=True)

        # If matches were found, add the DataFrame to the dictionary
        if not df_table_matches.empty:
            matches_dict[table_name] = df_table_matches

    return matches_dict



def process_client_table(db_path: str, client_table: str, client_key_field: str, update_clients: bool = True):
    """
    Process the client table to identify erroneous client codes, generate new unique codes, 
    save a mapping of old and new codes to an Excel file, and optionally update the client table 
    in the database.

    This function retrieves client data from the specified table, identifies erroneous codes 
    (e.g., those with a trailing letter), generates unique replacements, and stores the mapping 
    between old and new codes. The new codes are optionally updated in the database.

    Args:
        db_path (str): Path to the Access database file.
        client_table (str): The name of the table containing client data.
        client_key_field (str): The field in the table that contains the client codes.
        update_clients (bool, optional): If True, updates the client codes in the database with 
                                         the newly generated codes. Defaults to True.

    Returns:
        pd.DataFrame: A DataFrame containing the mapping of old and new client codes.
    """
    
    engine = get_dbaccess_connection(db_path)
    
    # Query to load the client data from the database
    query = f"SELECT * FROM {client_table}"
    df_clients = pd.read_sql(query, engine)

    # Filter erroneous codes that have a trailing letter
    df_clients['Has_Letter'] = df_clients[client_key_field].apply(has_trailing_letter)
    df_erroneous = df_clients[df_clients['Has_Letter'] == True]

    # Create the mapping table for old and new codes
    df_mappings = pd.DataFrame(columns=['OLD_CODE', 'NEW_CODE'])

    existing_codes = set(df_clients[client_key_field])  # Set of existing codes

    # Generate new codes
    for idx, row in df_erroneous.iterrows():
        old_code = row[client_key_field]
        new_code = generate_unique_code(old_code, existing_codes)
        existing_codes.add(new_code)  # Add new code to the existing set

        # Add old and new code to the mappings DataFrame
        new_row = pd.DataFrame([{'OLD_CODE': old_code, 'NEW_CODE': new_code}])
        df_mappings = pd.concat([df_mappings, new_row], ignore_index=True)
        
    # Conditionally update the client table in the database
    if update_clients:
        with engine.begin() as connection:
            for idx, row in df_mappings.iterrows():
                old_code = row['OLD_CODE']
                new_code = row['NEW_CODE']
                
                # SQL query to update the records directly
                query = sa.text(f"""
                    UPDATE {client_table}
                    SET {client_key_field} = :new_code
                    WHERE {client_key_field} = :old_code
                """)
                connection.execute(query, {'new_code': new_code, 'old_code': old_code})

    return df_mappings


def update_old_codes_in_db(matches_dict: dict, db_path: str, db_keys: dict) -> dict:
    """
    Updates the OLD_CODE values in the database tables with their corresponding NEW_CODE values 
    based on the matches found in the find_code_matches_in_db function. It also counts how many 
    records were updated for each table. All updates are handled within a transaction.

    Args:
        matches_dict (dict): Dictionary where keys are table names, and values are DataFrames 
                             with matches. The DataFrames must contain the key columns, 
                             FOUND_FIELD, FOUND_VALUE, and NEW_VALUE columns.
        db_path (str): Path to the Access database.
        db_keys (dict): Dictionary containing key fields for each table.

    Returns:
        dict: A dictionary where keys are table names and values are the count of updated records.
    """
    # Connect to the database
    engine = get_dbaccess_connection(db_path)

    # Dictionary to store the number of updated records for each table
    update_counts = {}

    # Begin a transaction
    with engine.begin() as connection:
        try:
            # Loop through each table in the matches_dict
            for table_name, df_matches in matches_dict.items():
                # Get the key columns for the current table
                key_columns = db_keys.get(table_name, [])

                # Initialize the count for this table
                update_counts[table_name] = 0

                # Loop through each row in the DataFrame to build the UPDATE query
                for idx, row in df_matches.iterrows():
                    found_value = row['FOUND_VALUE']
                    new_value = row['NEW_VALUE']
                    found_field = row['FOUND_FIELD']

                    # Build the WHERE clause using the key columns
                    where_clauses = []
                    for key_col in key_columns:
                        key_value = row[key_col]
                        where_clauses.append(f"{key_col} = :{key_col}")

                    # Join the WHERE clauses
                    where_clause = " AND ".join(where_clauses)

                    # Build the UPDATE query
                    update_query = sa.text(f"""
                    UPDATE {table_name}
                    SET {found_field} = :new_value
                    WHERE {found_field} = :found_value AND {where_clause}
                    """)

                    # print(f"Executing query: {update_query}")

                    # Prepare the parameter dictionary for the query
                    params = {'new_value': new_value, 'found_value': found_value}
                    for key_col in key_columns:
                        params[key_col] = row[key_col]

                    # Execute the UPDATE query and count successful updates
                    try:
                        result = connection.execute(update_query, params)
                        rows_updated = result.rowcount  # Get the number of rows affected
                        update_counts[table_name] += rows_updated  # Add to the counter
                    except Exception as e:
                        print(f"Error updating {table_name}: {e}")
                        raise  # Re-raise the exception to ensure the transaction is rolled back

        except Exception as e:
            print(f"Transaction failed: {e}")
            # Transaction will automatically roll back if any exception occurs

    return update_counts




